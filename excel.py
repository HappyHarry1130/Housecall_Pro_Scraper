import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

def write_to_excel(data, file_path, sheet_name="Data"):
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
        try:
            workbook = load_workbook(file_path)
        except FileNotFoundError:
            workbook = Workbook() 

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]  
        else:
            sheet = workbook.create_sheet(title=sheet_name)  

        headers = ["DATE", "Customer", "Link", "PART", "TOTAL SALE", "TOTAL"]

 
        header_fill = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")
        header_font = Font(bold=True, color="000000") 
        center_alignment = Alignment(horizontal='center', vertical='center') 

        if sheet.max_row == 1:
            sheet.append(headers)


            for cell in sheet[1]: 
                cell.fill = header_fill  
                cell.font = header_font 
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin')) 
                cell.alignment = center_alignment 


        sheet.append(data)

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):  
            for cell in row:
                if cell.row != 1:
                    cell.border = thin_border  
                    cell.alignment = center_alignment 

        column_widths = {
            'A': 15,  
            'B': 20,  
            'C': 45,  
            'D': 15,  
            'E': 15,  
            'F': 15   
        }

        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width  


        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")

